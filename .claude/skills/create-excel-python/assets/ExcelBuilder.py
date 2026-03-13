"""
Utility for writing styled Excel files.
"""
from pathlib import Path
import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelBuilder:
    """
    Writes a polars DataFrame to a styled Excel file with:
    - Table with auto-filters (via polars/xlsxwriter)
    - Frozen header row
    - Auto-fitted column widths based on actual cell content
    """

    MAX_COL_WIDTH = 60
    MIN_COL_WIDTH = 8
    PADDING = 2

    @classmethod
    def write(cls, df: pl.DataFrame, path: Path, sheet_name: str = "Sheet1") -> None:
        """Write DataFrame to Excel with filters, frozen header, and autofitted columns."""
        df.write_excel(str(path), worksheet=sheet_name)
        cls._post_process(path, sheet_name)

    @classmethod
    def _post_process(cls, path: Path, sheet_name: str) -> None:
        """Apply freeze panes and column autofit using openpyxl."""
        wb = load_workbook(str(path))
        ws = wb[sheet_name]

        # Autofit: measure max character length per column (header + data)
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            width = max(cls.MIN_COL_WIDTH, min(max_len + cls.PADDING, cls.MAX_COL_WIDTH))
            ws.column_dimensions[col_letter].width = width

        # Freeze header row so it stays visible when scrolling
        ws.freeze_panes = "A2"

        wb.save(str(path))
